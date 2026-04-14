"""Append tailored job metadata to the Excel job application tracker."""

from __future__ import annotations

import logging
from pathlib import Path

from openai_client import TrackerDetails

LOGGER = logging.getLogger(__name__)

EXPECTED_HEADERS = [
    "Company Name",
    "Role Title",
    "Location",
    "DS role type",
    "Alignment strength with my background",
    "Salary Range",
    "Job ID",
    "Date Applied",
    "Comments",
    "Status",
]


class TrackerUpdateError(RuntimeError):
    """Raised when the job tracker workbook cannot be updated safely."""


def append_tracker_row(*, tracker_path: str | Path, tracker_details: TrackerDetails) -> Path:
    """Append a row to the job application tracker workbook."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment dependency path
        raise TrackerUpdateError(
            "Updating the Excel tracker requires openpyxl. Install it with: ./.venv/bin/python -m pip install openpyxl"
        ) from exc

    tracker_path = Path(tracker_path).expanduser().resolve()
    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    if not tracker_path.exists():
        _create_tracker_workbook(tracker_path)

    workbook = load_workbook(tracker_path)
    worksheet = workbook[workbook.sheetnames[0]]
    headers = [worksheet.cell(row=1, column=index).value for index in range(1, worksheet.max_column + 1)]
    if headers != EXPECTED_HEADERS:
        normalized_headers = [header for header in headers if header is not None]
        if set(normalized_headers) != set(EXPECTED_HEADERS) or len(normalized_headers) != len(EXPECTED_HEADERS):
            raise TrackerUpdateError(
                f"Tracker headers do not match the expected schema in {tracker_path}."
            )

    if any(header is None for header in headers):
        raise TrackerUpdateError(
            f"Tracker headers do not match the expected schema in {tracker_path}."
        )

    values_by_header = {
        "Company Name": _normalize_tracker_value(tracker_details.company_name),
        "Role Title": _normalize_tracker_value(tracker_details.role_title),
        "Location": _normalize_tracker_value(tracker_details.location),
        "DS role type": _normalize_tracker_value(tracker_details.ds_role_type),
        "Alignment strength with my background": _normalize_tracker_value(
            tracker_details.alignment_strength_comment
        ),
        "Salary Range": _normalize_tracker_value(tracker_details.salary_range),
        "Job ID": _normalize_tracker_value(tracker_details.job_id),
        "Date Applied": "",
        "Comments": "",
        "Status": "",
    }

    row_values = [values_by_header[header] for header in headers]

    worksheet.append(row_values)
    workbook.save(tracker_path)
    LOGGER.info("Appended tracker row to %s", tracker_path)
    return tracker_path


def _normalize_tracker_value(value: str) -> str:
    value = (value or "").strip()
    return value if value else "unknown"


def _create_tracker_workbook(tracker_path: Path) -> None:
    """Create a new tracker workbook with the expected header row."""

    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - environment dependency path
        raise TrackerUpdateError(
            "Updating the Excel tracker requires openpyxl. Install it with: ./.venv/bin/python -m pip install openpyxl"
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Applications"
    worksheet.append(EXPECTED_HEADERS)
    workbook.save(tracker_path)
    LOGGER.info("Created new tracker workbook at %s", tracker_path)
